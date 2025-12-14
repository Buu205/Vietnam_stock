
import openpyxl
import pandas as pd
import json

FILE_PATH = '/Users/buuphan/Dev/Vietnam_dashboard/Brokerage Masterfile.xlsx'


def inspect_excel():
    print(f"Loading {FILE_PATH}...")
    try:

        wb = openpyxl.load_workbook(FILE_PATH, data_only=False)
        print("Sheet names:", wb.sheetnames)
        print("Defined names:", list(wb.defined_names.keys()))
        
        if "INPUT" in wb.defined_names:
             defn = wb.defined_names["INPUT"]
             print(f"INPUT defined name: {defn.value}")

        target_code = "MT.99"
        found_sheet = None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Search first 20 columns and 100 rows
            for row in ws.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and str(cell.value).strip() == target_code:
                        print(f"DTO Found {target_code} in sheet: {sheet_name} at {cell.coordinate}")
                        found_sheet = sheet_name
                        break
                if found_sheet: break
            if found_sheet: break
            

        if found_sheet:
            print(f"\n--- Inspecting {found_sheet} for Formulas ---")
            ws = wb[found_sheet]
            
            # Find specific rows
            targets = ["MT.99", "MT.100", "MT.101"]
            for row in ws.iter_rows(min_row=1, max_row=200):
                cell_val = str(row[0].value).strip()
                if cell_val in targets:
                    print(f"\nRow for {cell_val}:")
                    # Print first 5 columns to see formula
                    print([str(c.value) for c in row[:5]])


        if "Master" in wb.sheetnames:
            print("\n--- Extracting All MT Metrics from Master ---")
            ws = wb["Master"]
            mt_metrics = []
            for row in ws.iter_rows(min_row=1, max_row=500):
                val = str(row[0].value).strip()
                if val.startswith("MT."):
                    desc = str(row[1].value).strip()
                    mt_metrics.append((val, desc))
            
            print(f"Found {len(mt_metrics)} metrics.")
            for code, name in mt_metrics:
                print(f"{code} | {name}")

        if "INPUT" in wb.sheetnames:
            print("\n--- Inspecting INPUT Sheet for MT.100 ---")
            ws = wb["INPUT"]
            for row in ws.iter_rows(min_row=1, max_row=5000): # Check more rows
                if str(row[0].value).strip() == "MT.100":
                     print(f"Found MT.100 at row {row[0].row}")
                     # Print value and type of a few cells
                     print(f"Cell B{row[0].row} value: {row[1].value}")
                     # In openpyxl, if it's a formula, value is typically the formula string if data_only=False
                     # Let's check a few columns
                     print("Values/Formulas:", [str(c.value) for c in row[1:5]])
                     break

    except Exception as e:
        print(f"Error inspecting excel: {e}")

if __name__ == "__main__":
    inspect_excel()

